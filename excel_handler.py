import json
import re
from io import BytesIO

from anthropic import Anthropic
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_client: Anthropic | None = None
EXCEL_MODEL = "claude-sonnet-4-6"

# Xovia blue header + alternating row colours
_HEADER_HEX  = "185FA5"
_GREY_HEX    = "F5F5F5"
_WHITE_HEX   = "FFFFFF"

# Column names that get currency formatting
_CURRENCY_HEADERS = frozenset({
    "amount", "total", "cost", "price", "budget", "revenue",
    "salary", "income", "expense", "expenses", "profit", "loss",
    "subtotal", "tax",
})

_STRUCTURE_SYSTEM = """\
You are an Excel file architect for Xovia AI. Never mention Claude, Anthropic, or any AI provider.

When asked to create a spreadsheet, respond ONLY with a valid JSON object describing the workbook.
No markdown, no explanation, no preamble — raw JSON only.

Required format:
{
  "filename": "budget_tracker.xlsx",
  "sheets": [
    {
      "name": "How to Use",
      "type": "instructions",
      "rows": [
        ["Overview", "What this spreadsheet does and why it's useful — 2-3 sentences."],
        ["Sheet: Dashboard", "What to find here and what it updates automatically."],
        ["Sheet: Income", "Step-by-step: how to add a new income entry row by row."],
        ["Sheet: Expenses", "Step-by-step: how to record an expense and which columns to fill in."],
        ["Tips", "Bullet-point tips: save often, don't delete header rows, how totals update, etc."]
      ],
      "col_widths": [28, 90]
    },
    {
      "name": "Dashboard",
      "headers": ["Category", "Amount", "Percentage"],
      "rows": [["Income", 5000, "100%"], ["Expenses", 3200, "64%"]],
      "col_widths": [20, 15, 15]
    }
  ],
  "summary": "One sentence describing what was built.",
  "vba_code": "(optional) Include only if VBA automation would genuinely improve this spreadsheet. Omit the key entirely if not needed."
}

Rules:
- The FIRST sheet MUST always be "How to Use" with type "instructions"
- Write the How to Use sheet as if explaining to someone who has never used a spreadsheet
- Include a row for every other sheet explaining what it does and exactly how to use it
- After How to Use, create all logical data sheets (Dashboard, data entry sheets, Settings, etc.)
- Include realistic sample data in every data entry sheet

CRITICAL — FORMULAS:
- Dashboard and summary sheets MUST use Excel cross-sheet formulas, NOT hardcoded numbers.
  Example: to sum the Amount column from an Expenses sheet use "=SUM(Expenses!C:C)"
  Example: to count rows in a Sales Log sheet use "=COUNTA(Sales Log!A:A)-1"
  Example: net savings formula: "=SUM(Income!C:C)-SUM(Expenses!C:C)"
- Any cell that summarises or totals data from another sheet MUST be a formula string starting with =
- Data entry sheets (Income, Expenses, Sales Log, etc.) should have real sample rows with numeric values
- The Dashboard reads FROM those sheets via formulas — never duplicate numbers between sheets
- Percentages on the Dashboard should also be formulas e.g. "=B3/B2" not "64%"

- col_widths values are in Excel character units (typical: 15-25; use 90 for the instructions column)
- vba_code should be complete, working VBA — omit if not needed\
"""


def _get_client() -> Anthropic:
    global _client
    if _client is None:
        _client = Anthropic()
    return _client


def _get_structure(prompt: str) -> dict:
    """Ask Claude to describe the workbook as JSON and return the parsed dict."""
    response = _get_client().messages.create(
        model=EXCEL_MODEL,
        max_tokens=4096,
        system=_STRUCTURE_SYSTEM,
        messages=[{"role": "user", "content": prompt}],
    )
    text = response.content[0].text.strip()

    # Try direct parse first (model returned clean JSON)
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    # Slice from the first { to the last } — handles any fence/preamble wrapping
    start = text.find("{")
    end   = text.rfind("}") + 1
    if start >= 0 and end > start:
        return json.loads(text[start:end])

    raise ValueError(f"No JSON object found in model response: {text[:300]}")


def _build_instructions_sheet(ws, sheet_def: dict) -> None:
    """Render a How-to-Use / instructions sheet with wrapped text and taller rows."""
    rows   = sheet_def.get("rows", [])
    widths = sheet_def.get("col_widths", [28, 90])

    title_fill  = PatternFill(start_color=_HEADER_HEX, end_color=_HEADER_HEX, fill_type="solid")
    title_font  = Font(bold=True, color=_WHITE_HEX, size=13)
    label_font  = Font(bold=True, size=11)
    body_align  = Alignment(wrap_text=True, vertical="top")
    label_align = Alignment(wrap_text=True, vertical="top")

    # Title bar
    ws.row_dimensions[1].height = 30
    title_cell = ws.cell(row=1, column=1, value="How to Use This Spreadsheet")
    title_cell.fill      = title_fill
    title_cell.font      = title_font
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    # Span the title across both columns visually by making col B blank with same fill
    ws.cell(row=1, column=2).fill = title_fill

    grey_fill  = PatternFill(start_color=_GREY_HEX,  end_color=_GREY_HEX,  fill_type="solid")
    white_fill = PatternFill(start_color=_WHITE_HEX, end_color=_WHITE_HEX, fill_type="solid")

    for ri, row_data in enumerate(rows, 2):
        label = str(row_data[0]) if len(row_data) > 0 else ""
        body  = str(row_data[1]) if len(row_data) > 1 else ""

        fill = white_fill if ri % 2 == 0 else grey_fill
        ws.row_dimensions[ri].height = max(45, 15 * max(1, len(body) // 80))

        lbl_cell = ws.cell(row=ri, column=1, value=label)
        lbl_cell.font      = label_font
        lbl_cell.fill      = fill
        lbl_cell.alignment = label_align

        body_cell = ws.cell(row=ri, column=2, value=body)
        body_cell.fill      = fill
        body_cell.alignment = body_align

    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def _build_xlsx(structure: dict) -> bytes:
    """Build an .xlsx file from a structure dict and return raw bytes."""
    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    hdr_fill   = PatternFill(start_color=_HEADER_HEX, end_color=_HEADER_HEX, fill_type="solid")
    grey_fill  = PatternFill(start_color=_GREY_HEX,   end_color=_GREY_HEX,   fill_type="solid")
    white_fill = PatternFill(start_color=_WHITE_HEX,  end_color=_WHITE_HEX,  fill_type="solid")
    hdr_font   = Font(bold=True, color=_WHITE_HEX, size=11)
    centre     = Alignment(horizontal="center", vertical="center")

    for sheet_def in structure.get("sheets", []):
        name = str(sheet_def.get("name", "Sheet"))[:31]
        ws   = wb.create_sheet(title=name)

        # Instructions sheet — special layout
        if sheet_def.get("type") == "instructions" or name.lower() in ("how to use", "instructions", "readme"):
            _build_instructions_sheet(ws, sheet_def)
            continue

        headers = sheet_def.get("headers", [])
        rows    = sheet_def.get("rows", [])
        widths  = sheet_def.get("col_widths", [])

        ws.row_dimensions[1].height = 22

        # ── Header row ──────────────────────────────────────────────────────
        for ci, hdr in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=hdr)
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = centre

        # ── Data rows ───────────────────────────────────────────────────────
        for ri, row_data in enumerate(rows, 2):
            fill = white_fill if ri % 2 == 0 else grey_fill
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill = fill
                if ci <= len(headers):
                    if (
                        any(kw in headers[ci - 1].lower() for kw in _CURRENCY_HEADERS)
                        and isinstance(val, (int, float))
                    ):
                        cell.number_format = '$#,##0.00'

        # ── Column widths ───────────────────────────────────────────────────
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        for ci in range(len(widths) + 1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 18

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_excel(prompt: str) -> tuple[bytes | None, str, str, str | None]:
    """
    Generate an Excel workbook from a user prompt.
    Returns (excel_bytes, filename, chat_response, error_key).
    error_key is None on success.
    """
    try:
        structure = _get_structure(prompt)
    except Exception as e:
        return None, "", f"I couldn't generate the Excel structure. ({e})", "parse_error"

    try:
        excel_bytes = _build_xlsx(structure)
    except Exception as e:
        return None, "", f"I couldn't build the Excel file. ({e})", "build_error"

    filename = str(structure.get("filename", "spreadsheet.xlsx"))
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"

    summary  = structure.get("summary", "Your spreadsheet is ready.")
    vba_code = structure.get("vba_code", "").strip()

    lines = [f"✅ {summary}", "", "Your file is ready — click **Download as .xlsx** below."]

    if vba_code:
        lines += [
            "",
            "---",
            "### VBA Automation Code",
            "Paste this into the Visual Basic Editor (**Alt+F11** in Excel) to add automation:",
            f"```vba\n{vba_code}\n```",
            "",
            "> **Note:** VBA runs inside Excel after you paste it — it is not embedded in the "
            "downloaded .xlsx file.",
        ]

    return excel_bytes, filename, "\n".join(lines), None


# ---------------------------------------------------------------------------
# Standalone test — run with: python excel_handler.py
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import os
    import tomllib

    with open(".streamlit/secrets.toml", "rb") as f:
        _s = tomllib.load(f)
    os.environ["ANTHROPIC_API_KEY"] = _s["anthropic"]["api_key"]

    prompts = [
        "Can you build me a budget tracker Excel file?",
        "Create a simple sales tracker spreadsheet with columns for date, product, quantity, and revenue",
    ]
    for i, p in enumerate(prompts, 1):
        print(f"\n[{i}] {p}")
        xb, fn, resp, err = generate_excel(p)
        if err:
            print(f"    ERROR: {err}")
        else:
            with open(f"test_excel_{i}.xlsx", "wb") as fh:
                fh.write(xb)
            print(f"    Saved: test_excel_{i}.xlsx  ({len(xb):,} bytes)")
            print(f"    Response:\n{resp[:200]}...")
